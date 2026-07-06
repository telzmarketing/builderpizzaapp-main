from __future__ import annotations

import csv
import json
import random
import re
import uuid
import unicodedata
from datetime import datetime, time, timedelta, timezone
from io import BytesIO, StringIO
from zoneinfo import ZoneInfo

from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session, joinedload, selectinload

from backend.models.customer import Address
from backend.models.order import Order, OrderItem, OrderStatus
from backend.models.payment import Payment, PaymentStatus
from backend.models.product import Product
from backend.models.store_notification import (
    StoreNotification,
    StoreNotificationCaptured,
    StoreNotificationDay,
    StoreNotificationImpression,
    StoreNotificationSettings,
)
from backend.schemas.store_notification import (
    StoreNotificationCreate,
    StoreNotificationPreviewIn,
    StoreNotificationSettingsIn,
    StoreNotificationUpdate,
)
from backend.services.store_operation_service import StoreOperationService


ALLOWED_PAGES = {"home", "cardapio", "product", "cart"}
DEFAULT_TEMPLATE = "{nome} — {bairro} comprou hoje {produto}"
FIXED_DISPLAY_TEMPLATE = "{name} — {neighborhood} comprou hoje {product}"
PAID_ORDER_STATUSES = {
    OrderStatus.paid,
    OrderStatus.pago,
    OrderStatus.preparing,
    OrderStatus.ready_for_pickup,
    OrderStatus.on_the_way,
    OrderStatus.delivered,
}
PAID_PAYMENT_STATUSES = {PaymentStatus.approved, PaymentStatus.paid}
PRIORITY_SCORE = {"low": 1, "medium": 2, "high": 3}
CAPTURE_LOOKBACK_HOURS = 72
MAX_IMPORT_ROWS = 1000
IMPORT_HEADER_ALIASES = {
    "nome": "display_name",
    "nomedocomprador": "display_name",
    "comprador": "display_name",
    "cliente": "display_name",
    "buyer": "display_name",
    "buyername": "display_name",
    "displayname": "display_name",
    "produtoid": "product_id",
    "idproduto": "product_id",
    "productid": "product_id",
    "produto": "product_name",
    "nomeproduto": "product_name",
    "produtonome": "product_name",
    "product": "product_name",
    "productname": "product_name",
    "bairro": "neighborhood",
    "neighborhood": "neighborhood",
    "regiao": "neighborhood",
    "minutos": "purchase_minutes_ago",
    "minutoscompra": "purchase_minutes_ago",
    "compradoha": "purchase_minutes_ago",
    "purchaseminutesago": "purchase_minutes_ago",
}


class StoreNotificationService:
    def __init__(self, db: Session):
        self._db = db

    def get_settings(self) -> StoreNotificationSettings:
        settings = (
            self._db.query(StoreNotificationSettings)
            .filter(StoreNotificationSettings.id == "default")
            .first()
        )
        if settings:
            return settings
        settings = StoreNotificationSettings(id="default")
        self._db.add(settings)
        self._db.commit()
        self._db.refresh(settings)
        return settings

    def update_settings(self, payload: StoreNotificationSettingsIn) -> StoreNotificationSettings:
        settings = self.get_settings()
        values = payload.model_dump()
        for key, value in values.items():
            if key == "allowed_pages":
                value = json.dumps(value)
            setattr(settings, key, value)
        settings.updated_at = datetime.now(timezone.utc)
        self._db.commit()
        self._db.refresh(settings)
        return settings

    def list_notifications(self) -> list[dict]:
        notifications = (
            self._db.query(StoreNotification)
            .options(selectinload(StoreNotification.days), joinedload(StoreNotification.product))
            .order_by(StoreNotification.created_at.desc())
            .all()
        )
        return [self.serialize_notification(item) for item in notifications]

    def summary(self) -> dict:
        active = (
            self._db.query(func.count(StoreNotification.id))
            .filter(StoreNotification.status == "active")
            .scalar()
            or 0
        )
        manual = (
            self._db.query(func.count(StoreNotification.id))
            .filter(StoreNotification.type.in_(["manual", "fomento"]))
            .scalar()
            or 0
        )
        real_impressions = (
            self._db.query(func.count(StoreNotificationImpression.id))
            .filter(StoreNotificationImpression.source_type == "real")
            .scalar()
            or 0
        )
        total_impressions = self._db.query(func.count(StoreNotificationImpression.id)).scalar() or 0
        return {
            "active_notifications": active,
            "manual_notifications": manual,
            "real_impressions": real_impressions,
            "total_impressions": total_impressions,
        }

    def create_notification(self, payload: StoreNotificationCreate, source_customer_id: str | None = None) -> StoreNotification:
        self._ensure_product(payload.product_id)
        values = payload.model_dump(exclude={"weekdays"})
        notification = StoreNotification(
            id=f"sn-{uuid.uuid4().hex[:10]}",
            source_customer_id=source_customer_id,
            **values,
        )
        self._db.add(notification)
        self._replace_days(notification, payload.weekdays)
        self._db.commit()
        self._db.refresh(notification)
        return notification

    def update_notification(self, notification_id: str, payload: StoreNotificationUpdate) -> StoreNotification:
        notification = self._notification(notification_id)
        values = payload.model_dump(exclude_unset=True, exclude={"weekdays"})
        if "product_id" in values and values["product_id"]:
            self._ensure_product(values["product_id"])
        for key, value in values.items():
            setattr(notification, key, value)
        if payload.weekdays is not None:
            self._replace_days(notification, payload.weekdays)
        notification.updated_at = datetime.now(timezone.utc)
        self._db.commit()
        self._db.refresh(notification)
        return notification

    def duplicate_notification(self, notification_id: str) -> StoreNotification:
        notification = self._notification(notification_id)
        copy = StoreNotification(
            id=f"sn-{uuid.uuid4().hex[:10]}",
            type=notification.type,
            status="paused",
            internal_name=f"{notification.internal_name} - copia",
            display_name=notification.display_name,
            product_id=notification.product_id,
            neighborhood=notification.neighborhood,
            template_text=notification.template_text,
            priority=notification.priority,
            weight=notification.weight,
            display_seconds=notification.display_seconds,
            purchase_minutes_ago=notification.purchase_minutes_ago or 12,
            clear_after_view=bool(notification.clear_after_view),
            start_time=notification.start_time,
            end_time=notification.end_time,
            start_date=notification.start_date,
            end_date=notification.end_date,
            source_customer_id=notification.source_customer_id,
        )
        self._db.add(copy)
        self._replace_days(copy, [day.weekday for day in notification.days])
        self._db.commit()
        self._db.refresh(copy)
        return copy

    def set_status(self, notification_id: str, status: str) -> StoreNotification:
        if status not in {"active", "paused"}:
            raise ValueError("Status invalido.")
        notification = self._notification(notification_id)
        notification.status = status
        notification.updated_at = datetime.now(timezone.utc)
        self._db.commit()
        self._db.refresh(notification)
        return notification

    def delete_notification(self, notification_id: str) -> None:
        notification = self._notification(notification_id)
        self._db.delete(notification)
        self._db.commit()

    def import_notifications_file(self, filename: str, contents: bytes) -> dict:
        rows = self._parse_import_file(filename, contents)
        if not rows:
            raise ValueError("Arquivo sem linhas para importar.")
        if len(rows) > MAX_IMPORT_ROWS:
            raise ValueError(f"Importacao limitada a {MAX_IMPORT_ROWS} linhas por arquivo.")

        created = []
        errors = []
        for row_number, row in rows:
            try:
                payload = self._payload_from_import_row(row)
                notification = self.create_notification(payload)
                created.append(self.serialize_notification(notification))
            except (LookupError, ValueError) as exc:
                errors.append({"row": row_number, "message": str(exc)})

        return {
            "created_count": len(created),
            "skipped_count": len(errors),
            "errors": errors,
            "notifications": created,
        }

    def _parse_import_file(self, filename: str, contents: bytes) -> list[tuple[int, dict[str, str]]]:
        name = (filename or "").lower().strip()
        if name.endswith(".csv"):
            return self._parse_csv_import(contents)
        if name.endswith(".xlsx"):
            return self._parse_xlsx_import(contents)
        raise ValueError("Envie um arquivo .csv ou .xlsx.")

    def _parse_csv_import(self, contents: bytes) -> list[tuple[int, dict[str, str]]]:
        text = None
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                text = contents.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("Nao foi possivel ler o CSV. Use UTF-8 ou Latin-1.")

        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(StringIO(text), dialect=dialect)
        if not reader.fieldnames:
            raise ValueError("CSV sem cabecalho.")
        return self._normalize_import_rows(
            (index, row) for index, row in enumerate(reader, start=2)
        )

    def _parse_xlsx_import(self, contents: bytes) -> list[tuple[int, dict[str, str]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Leitura de Excel indisponivel. Instale openpyxl no backend.") from exc

        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        header_names = [str(value or "").strip() for value in headers]
        raw_rows = []
        for row_number, values in enumerate(rows, start=2):
            raw_rows.append((row_number, dict(zip(header_names, values))))
        return self._normalize_import_rows(raw_rows)

    def _normalize_import_rows(self, rows) -> list[tuple[int, dict[str, str]]]:
        normalized_rows = []
        for row_number, row in rows:
            normalized = {}
            for key, value in row.items():
                mapped_key = IMPORT_HEADER_ALIASES.get(self._normalize_import_key(key))
                if not mapped_key:
                    continue
                normalized[mapped_key] = self._safe_text(str(value)) if value is not None else None
            if any(normalized.values()):
                normalized_rows.append((row_number, normalized))
        return normalized_rows

    def _payload_from_import_row(self, row: dict[str, str]) -> StoreNotificationCreate:
        buyer_name = self._safe_text(row.get("display_name"))
        if not buyer_name:
            raise ValueError("Informe o nome do comprador.")

        product_id = self._product_id_from_import_row(row)
        neighborhood = self._safe_text(row.get("neighborhood"))
        if not neighborhood:
            raise ValueError("Informe o bairro do comprador.")

        purchase_minutes = self._parse_purchase_minutes(row.get("purchase_minutes_ago"))
        product = self._db.query(Product).filter(Product.id == product_id).first()
        product_name = self._product_display_name(product, product.name if product else None) or "Produto"
        display_name = self._first_name(buyer_name)

        return StoreNotificationCreate(
            type="manual",
            status="active",
            internal_name=f"Importado: {display_name} - {product_name}"[:200],
            display_name=display_name,
            product_id=product_id,
            neighborhood=neighborhood,
            template_text=DEFAULT_TEMPLATE,
            priority="medium",
            weight=1,
            display_seconds=7,
            purchase_minutes_ago=purchase_minutes,
            clear_after_view=False,
            start_time=time(18, 0),
            end_time=time(23, 30),
            start_date=None,
            end_date=None,
            weekdays=[0, 1, 2, 3, 4, 5, 6],
        )

    def _product_id_from_import_row(self, row: dict[str, str]) -> str:
        product_id = self._safe_text(row.get("product_id"))
        if product_id:
            self._ensure_product(product_id)
            return product_id

        product_name = self._safe_text(row.get("product_name"))
        if not product_name:
            raise ValueError("Informe o produto ou product_id.")

        target = self._normalize_import_key(product_name.removeprefix("Pizza "))
        products = self._db.query(Product).all()
        for product in products:
            names = [product.name, self._product_display_name(product, product.name)]
            for name in names:
                normalized = self._normalize_import_key(name)
                normalized_without_pizza = self._normalize_import_key((name or "").removeprefix("Pizza "))
                if target in {normalized, normalized_without_pizza}:
                    return product.id
        raise ValueError(f"Produto nao encontrado: {product_name}.")

    def _parse_purchase_minutes(self, value: str | None) -> int:
        if not value:
            return 14
        try:
            minutes = int(float(str(value).replace(",", ".")))
        except ValueError as exc:
            raise ValueError("Minutos da compra deve ser numero inteiro positivo.") from exc
        if minutes <= 0 or minutes > 1440:
            raise ValueError("Minutos da compra deve ficar entre 1 e 1440.")
        return minutes

    def _normalize_import_key(self, value: str | None) -> str:
        normalized = unicodedata.normalize("NFKD", value or "")
        ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
        return re.sub(r"[^a-z0-9]+", "", ascii_value.lower())

    def preview(self, payload: StoreNotificationPreviewIn) -> dict:
        product_name = self._safe_text(payload.product_name) or "Produto"
        message = self._render_notification_message(
            name=self._safe_text(payload.display_name) or "Cliente",
            neighborhood=self._safe_text(payload.neighborhood),
            product=product_name,
        )
        return {"message": message}

    def next_notification(
        self,
        page: str = "home",
        customer_id: str | None = None,
        anonymous_session_id: str | None = None,
        seen_ids: list[str] | None = None,
    ) -> dict:
        settings = self.get_settings()
        next_delay = self._next_delay(settings)
        initial_delay = max(1, settings.initial_delay_seconds or 5)
        page = self._normalize_page(page)

        empty = {
            "notification": None,
            "next_delay_seconds": next_delay,
            "initial_delay_seconds": initial_delay,
        }

        if not settings.enabled or page not in self._settings_pages(settings):
            return empty

        if settings.only_during_store_hours:
            try:
                if not StoreOperationService(self._db).get_status()["is_open"]:
                    return empty
            except Exception:
                return empty

        last = self._last_impression()
        candidates = self._manual_candidates(
            settings,
            last,
            customer_id=customer_id,
            anonymous_session_id=anonymous_session_id,
            seen_ids=seen_ids or [],
        )

        if not candidates:
            return empty

        real_candidates = [item for item in candidates if item.get("source_type") == "real"]
        selected = self._weighted_manual(real_candidates or candidates)
        return {
            "notification": selected,
            "next_delay_seconds": next_delay,
            "initial_delay_seconds": initial_delay,
        }

    # ── Display impressions ──────────────────────────────────────────────────

    def record_impression(
        self,
        notification_id: str,
        *,
        page: str = "home",
        customer_id: str | None = None,
        anonymous_session_id: str | None = None,
    ) -> None:
        notification = self._notification(notification_id)
        if self._was_already_displayed(notification.id, customer_id, anonymous_session_id):
            return

        page = self._normalize_page(page)
        source_type = "real" if notification.source_customer_id else "manual"
        impression = StoreNotificationImpression(
            id=f"sni-{uuid.uuid4().hex[:12]}",
            notification_id=notification.id,
            source_type=source_type,
            product_id=notification.product_id,
            neighborhood=notification.neighborhood,
            page=page,
            customer_id=customer_id,
            anonymous_session_id=anonymous_session_id,
            notification_type=source_type,
            displayed_at=datetime.now(timezone.utc),
        )
        self._db.add(impression)
        if notification.clear_after_view:
            notification.status = "paused"
            notification.updated_at = datetime.now(timezone.utc)
        self._db.commit()

    # ── Captured notifications ────────────────────────────────────────────────

    def list_captured(self) -> list[dict]:
        settings = self.get_settings()
        if settings.real_orders_enabled:
            self._sync_captured_from_orders()
        items = (
            self._db.query(StoreNotificationCaptured)
            .order_by(StoreNotificationCaptured.created_at.desc())
            .all()
        )
        return [self._serialize_captured(item) for item in items]

    def discard_captured(self, captured_id: str) -> None:
        captured = self._captured(captured_id)
        captured.status = "discarded"
        self._db.commit()

    def activate_captured(self, captured_id: str, payload: StoreNotificationCreate) -> dict:
        captured = self._captured(captured_id)
        if captured.status == "activated":
            raise ValueError("Notificacao ja foi ativada anteriormente.")
        notification = self.create_notification(payload, source_customer_id=captured.customer_id)
        captured.status = "activated"
        self._db.commit()
        self._db.refresh(notification)
        return self.serialize_notification(notification)

    def _sync_captured_from_orders(self) -> None:
        cutoff = datetime.now(timezone.utc) - timedelta(hours=CAPTURE_LOOKBACK_HOURS)
        orders = (
            self._db.query(Order)
            .options(
                joinedload(Order.customer),
                joinedload(Order.address),
                joinedload(Order.payment),
                selectinload(Order.items).selectinload(OrderItem.flavors),
            )
            .outerjoin(Payment, Payment.order_id == Order.id)
            .filter(Order.created_at >= cutoff)
            .filter(
                or_(
                    Order.status.in_(list(PAID_ORDER_STATUSES)),
                    Order.paid_at.is_not(None),
                    Payment.status.in_(list(PAID_PAYMENT_STATUSES)),
                )
            )
            .filter(~Order.status.in_([OrderStatus.cancelled, OrderStatus.refunded]))
            .order_by(Order.created_at.desc())
            .limit(100)
            .all()
        )
        if not orders:
            return

        existing_order_ids = {
            row[0]
            for row in self._db.query(StoreNotificationCaptured.order_id)
            .filter(StoreNotificationCaptured.order_id.in_([o.id for o in orders]))
            .all()
        }
        product_lookup = self._product_lookup(orders)
        for order in orders:
            if order.id in existing_order_ids:
                continue
            product_id, product_name, product_image = self._main_product(order, product_lookup)
            if not product_name:
                continue
            neighborhood = self._order_neighborhood(order)
            buyer_name = self._first_name(
                order.customer.name if order.customer else order.delivery_name
            )
            captured = StoreNotificationCaptured(
                id=f"snc-{uuid.uuid4().hex[:10]}",
                order_id=order.id,
                customer_id=order.customer_id,
                product_id=product_id,
                product_name=product_name,
                product_image=product_image,
                neighborhood=neighborhood,
                buyer_name=buyer_name,
                order_time=order.paid_at or order.created_at,
            )
            self._db.add(captured)
        try:
            self._db.commit()
        except Exception:
            self._db.rollback()

    def _serialize_captured(self, item: StoreNotificationCaptured) -> dict:
        return {
            "id": item.id,
            "order_id": item.order_id,
            "customer_id": item.customer_id,
            "product_id": item.product_id,
            "product_name": item.product_name,
            "product_image": item.product_image,
            "neighborhood": item.neighborhood,
            "buyer_name": item.buyer_name,
            "order_time": item.order_time,
            "status": item.status,
            "created_at": item.created_at,
        }

    def _captured(self, captured_id: str) -> StoreNotificationCaptured:
        item = (
            self._db.query(StoreNotificationCaptured)
            .filter(StoreNotificationCaptured.id == captured_id)
            .first()
        )
        if not item:
            raise LookupError("Notificacao capturada nao encontrada.")
        return item

    # ── Settings serialization ────────────────────────────────────────────────

    def serialize_settings(self, settings: StoreNotificationSettings) -> dict:
        return {
            "id": settings.id,
            "enabled": bool(settings.enabled),
            "real_orders_enabled": bool(settings.real_orders_enabled),
            "real_percentage": settings.real_percentage,
            "manual_percentage": settings.manual_percentage,
            "initial_delay_seconds": settings.initial_delay_seconds,
            "min_delay_seconds": settings.min_delay_seconds,
            "max_delay_seconds": settings.max_delay_seconds,
            "default_display_seconds": settings.default_display_seconds,
            "prevent_same_product_sequence": bool(settings.prevent_same_product_sequence),
            "prevent_same_neighborhood_sequence": bool(settings.prevent_same_neighborhood_sequence),
            "only_during_store_hours": bool(settings.only_during_store_hours),
            "allowed_pages": self._settings_pages(settings),
            "created_at": settings.created_at,
            "updated_at": settings.updated_at,
        }

    def serialize_notification(self, notification: StoreNotification) -> dict:
        last_displayed_at = (
            self._db.query(func.max(StoreNotificationImpression.displayed_at))
            .filter(StoreNotificationImpression.notification_id == notification.id)
            .scalar()
        )
        impressions_count = (
            self._db.query(func.count(StoreNotificationImpression.id))
            .filter(StoreNotificationImpression.notification_id == notification.id)
            .scalar()
            or 0
        )
        return {
            "id": notification.id,
            "type": notification.type,
            "status": notification.status,
            "internal_name": notification.internal_name,
            "display_name": notification.display_name,
            "product_id": notification.product_id,
            "product_name": notification.product.name if notification.product else None,
            "product_icon": notification.product.icon if notification.product else None,
            "neighborhood": notification.neighborhood,
            "template_text": notification.template_text,
            "priority": notification.priority,
            "weight": notification.weight,
            "display_seconds": notification.display_seconds,
            "purchase_minutes_ago": notification.purchase_minutes_ago or 12,
            "clear_after_view": bool(notification.clear_after_view),
            "start_time": notification.start_time,
            "end_time": notification.end_time,
            "start_date": notification.start_date,
            "end_date": notification.end_date,
            "weekdays": [day.weekday for day in notification.days],
            "impressions_count": impressions_count,
            "last_displayed_at": last_displayed_at,
            "created_at": notification.created_at,
            "updated_at": notification.updated_at,
        }

    # ── Candidate selection ───────────────────────────────────────────────────

    def _manual_candidates(
        self,
        settings: StoreNotificationSettings,
        last: StoreNotificationImpression | None,
        customer_id: str | None = None,
        anonymous_session_id: str | None = None,
        seen_ids: list[str] | None = None,
    ) -> list[dict]:
        current = self._local_now()
        seen_set = set(seen_ids or [])
        notifications = (
            self._db.query(StoreNotification)
            .options(selectinload(StoreNotification.days), joinedload(StoreNotification.product))
            .filter(StoreNotification.status == "active")
            .filter(StoreNotification.type.in_(["manual", "fomento"]))
            .all()
        )
        candidates = []
        for item in notifications:
            if item.id in seen_set:
                continue
            if customer_id and item.source_customer_id and item.source_customer_id == customer_id:
                continue
            if self._was_already_displayed(item.id, customer_id, anonymous_session_id):
                continue
            if not self._manual_is_eligible(item, current):
                continue
            product_name = self._product_display_name(item.product, "Produto")
            purchase_minutes = int(item.purchase_minutes_ago or 0)
            if not self._has_complete_display_data(item, product_name):
                continue
            source_type = "real" if item.source_customer_id else "manual"
            candidates.append({
                "source_type": source_type,
                "notification_id": item.id,
                "product_id": item.product_id,
                "product_name": product_name,
                "product_image": item.product.icon if item.product else None,
                "neighborhood": self._safe_text(item.neighborhood),
                "message": self._render_notification_message(
                    name=self._safe_text(item.display_name) or "Cliente",
                    neighborhood=self._safe_text(item.neighborhood),
                    product=product_name,
                ),
                "display_seconds": item.display_seconds or settings.default_display_seconds,
                "purchase_minutes_ago": purchase_minutes,
                "_product_id": item.product_id,
                "_neighborhood": item.neighborhood,
                "_score": max(1, item.weight or 1) * PRIORITY_SCORE.get(item.priority, 2),
            })
        sequenced = [
            item for item in candidates
            if self._passes_sequence_rules(settings, last, item.get("_product_id"), item.get("_neighborhood"))
        ]
        if sequenced:
            candidates = sequenced
        for item in candidates:
            item.pop("_product_id", None)
            item.pop("_neighborhood", None)
        return candidates

    def _weighted_manual(self, candidates: list[dict]) -> dict:
        total = sum(item["_score"] for item in candidates)
        cursor = random.randint(1, max(1, total))
        running = 0
        for item in sorted(candidates, key=lambda candidate: candidate["_score"], reverse=True):
            running += item["_score"]
            if cursor <= running:
                item.pop("_score", None)
                return item
        candidates[0].pop("_score", None)
        return candidates[0]

    def _was_already_displayed(
        self,
        notification_id: str,
        customer_id: str | None,
        anonymous_session_id: str | None,
    ) -> bool:
        if not customer_id and not anonymous_session_id:
            return False
        query = self._db.query(StoreNotificationImpression.id).filter(
            StoreNotificationImpression.notification_id == notification_id
        )
        identity_filters = []
        if customer_id:
            identity_filters.append(StoreNotificationImpression.customer_id == customer_id)
        if anonymous_session_id:
            identity_filters.append(StoreNotificationImpression.anonymous_session_id == anonymous_session_id)
        return query.filter(or_(*identity_filters)).first() is not None

    def _has_complete_display_data(
        self,
        notification: StoreNotification,
        product_name: str | None,
    ) -> bool:
        return all([
            self._safe_text(notification.display_name),
            self._safe_text(product_name),
            self._safe_text(notification.neighborhood),
        ])

    def _manual_is_eligible(self, notification: StoreNotification, current: datetime) -> bool:
        if current.weekday() not in [day.weekday for day in notification.days]:
            return False
        today = current.date()
        if notification.start_date and notification.start_date > today:
            return False
        if notification.end_date and notification.end_date < today:
            return False
        now_time = current.time()
        if notification.start_time <= notification.end_time:
            return notification.start_time <= now_time <= notification.end_time
        return now_time >= notification.start_time or now_time <= notification.end_time

    def _passes_sequence_rules(
        self,
        settings: StoreNotificationSettings,
        last: StoreNotificationImpression | None,
        product_id: str | None,
        neighborhood: str | None,
    ) -> bool:
        if not last:
            return True
        if settings.prevent_same_product_sequence and product_id and last.product_id == product_id:
            return False
        if (
            settings.prevent_same_neighborhood_sequence
            and neighborhood
            and last.neighborhood
            and last.neighborhood.lower() == neighborhood.lower()
        ):
            return False
        return True

    def _last_impression(self) -> StoreNotificationImpression | None:
        return (
            self._db.query(StoreNotificationImpression)
            .order_by(StoreNotificationImpression.displayed_at.desc())
            .first()
        )

    def _replace_days(self, notification: StoreNotification, weekdays: list[int]) -> None:
        if notification.id:
            self._db.query(StoreNotificationDay).filter(
                StoreNotificationDay.notification_id == notification.id
            ).delete(synchronize_session=False)
        for weekday in sorted(set(weekdays)):
            self._db.add(StoreNotificationDay(
                id=f"snd-{uuid.uuid4().hex[:10]}",
                notification_id=notification.id,
                weekday=weekday,
            ))

    def _notification(self, notification_id: str) -> StoreNotification:
        notification = (
            self._db.query(StoreNotification)
            .options(selectinload(StoreNotification.days), joinedload(StoreNotification.product))
            .filter(StoreNotification.id == notification_id)
            .first()
        )
        if not notification:
            raise LookupError("Notificacao nao encontrada.")
        return notification

    def _ensure_product(self, product_id: str) -> None:
        exists = self._db.query(Product.id).filter(Product.id == product_id).first()
        if not exists:
            raise LookupError("Produto vinculado nao encontrado.")

    def _settings_pages(self, settings: StoreNotificationSettings) -> list[str]:
        try:
            pages = json.loads(settings.allowed_pages or "[]")
        except Exception:
            pages = []
        clean = [page for page in pages if page in ALLOWED_PAGES]
        return clean or ["home", "cardapio", "product", "cart"]

    def _normalize_page(self, page: str) -> str:
        page = (page or "home").strip().lower()
        aliases = {
            "/": "home",
            "catalog": "cardapio",
            "cardápio": "cardapio",
            "produto": "product",
            "carrinho": "cart",
        }
        return aliases.get(page, page if page in ALLOWED_PAGES else "home")

    def _next_delay(self, settings: StoreNotificationSettings) -> int:
        min_delay = max(1, settings.min_delay_seconds or 45)
        max_delay = max(min_delay, settings.max_delay_seconds or min_delay)
        return random.randint(min_delay, max_delay)

    def _local_now(self) -> datetime:
        return datetime.now(ZoneInfo("America/Sao_Paulo"))

    def _product_lookup(self, orders: list[Order]) -> dict[str, Product]:
        product_ids: set[str] = set()
        for order in orders:
            for item in order.items:
                product_ids.add(item.product_id)
                for flavor in item.flavors:
                    product_ids.add(flavor.product_id)
        if not product_ids:
            return {}
        products = self._db.query(Product).filter(Product.id.in_(product_ids)).all()
        return {product.id: product for product in products}

    def _main_product(self, order: Order, product_lookup: dict[str, Product]) -> tuple[str | None, str | None, str | None]:
        if not order.items:
            return None, None, None
        item = sorted(order.items, key=lambda candidate: candidate.total_price or 0, reverse=True)[0]
        if item.flavors:
            flavor = sorted(item.flavors, key=lambda candidate: candidate.position or 0)[0]
            product = product_lookup.get(flavor.product_id)
            return flavor.product_id, self._product_display_name(product, flavor.flavor_name), product.icon if product else None
        product = product_lookup.get(item.product_id)
        return item.product_id, self._product_display_name(product, product.name if product else None), product.icon if product else None

    def _order_neighborhood(self, order: Order) -> str | None:
        if order.address and order.address.neighborhood:
            return self._safe_text(order.address.neighborhood)
        if order.customer_id:
            address = (
                self._db.query(Address)
                .filter(Address.customer_id == order.customer_id)
                .order_by(Address.is_default.desc(), Address.created_at.desc())
                .first()
            )
            if address and address.neighborhood:
                return self._safe_text(address.neighborhood)
        return None

    def _render_template(
        self,
        template: str,
        *,
        name: str,
        product: str,
        product_with_article: str,
        neighborhood: str | None,
        relative_time: str,
    ) -> str:
        return self._render_notification_message(
            name=name,
            neighborhood=neighborhood,
            product=product,
        )

    def _render_notification_message(
        self,
        *,
        name: str,
        neighborhood: str | None,
        product: str,
    ) -> str:
        clean_name = self._safe_text(name) or "Cliente"
        clean_neighborhood = self._safe_text(neighborhood) or "bairro"
        clean_product = self._safe_text(product) or "Produto"
        return FIXED_DISPLAY_TEMPLATE.format(
            name=clean_name,
            neighborhood=clean_neighborhood,
            product=clean_product,
        )

    def _product_with_article(self, product_name: str | None) -> str:
        product = self._safe_text(product_name) or "produto"
        normalized = product.lower()
        if normalized.startswith(("pizza", "bebida", "agua", "água", "coca", "fanta", "sprite")):
            return f"uma {product}"
        return f"um {product}"

    def _product_display_name(self, product: Product | None, fallback: str | None) -> str | None:
        name = self._safe_text(product.name if product else fallback)
        if not name:
            return None
        if name.lower().startswith("pizza"):
            return name
        product_type = (product.product_type or "").lower() if product else ""
        category = " ".join(filter(None, [product.category, product.subcategory])).lower() if product else ""
        if product_type == "pizza" or "pizza" in category:
            return f"Pizza {name}"
        return name

    def _first_name(self, value: str | None) -> str:
        clean = self._safe_text(value)
        if not clean:
            return "Cliente"
        return clean.split()[0][:40]

    def _safe_text(self, value: str | None) -> str | None:
        if not value:
            return None
        clean = re.sub(r"[\r\n\t]+", " ", str(value)).strip()
        clean = re.sub(r"\s+", " ", clean)
        return clean[:120] or None
